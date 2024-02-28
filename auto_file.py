from openpyxl import load_workbook
import os
import shutil
import time
dest_file = "C:\\Users\\wjf\\Desktop\\cs"
xml_muban = "C:\\Users\\wjf\\Desktop\\安全厂家速查20240221.xlsx"

gongsi_file = "C:\\Users\\wjf\\Desktop\\安全公司梳理"

wb = load_workbook(xml_muban)
for i, index in enumerate(range(14)):
    sheet = wb.worksheets[i + 2]
    # print(wb.sheetnames[i+2])
    print("开始" + "----" + wb.sheetnames[i+2])

    for row in range(sheet.max_column):
        time.sleep(1)
        line = 1
        while True:
            line = line + 1
            cell = sheet.cell(line, row + 1)
            if cell.value is None:
                break
            else:
                lingshi_path = os.path.join(dest_file, wb.sheetnames[i + 2], sheet.cell(1, row + 1).value)
                if not os.path.exists(lingshi_path):
                    os.makedirs(lingshi_path)
                # print(lingshi_path)

                with open('error_log.txt', mode='at') as f:
                    if not os.path.exists(os.path.join(gongsi_file, cell.value + ".docx")):
                        f.write("\n" + cell.value)

                # print(os.path.join(gongsi_file, cell.value + ".docx"))
                # print(os.path.exists(os.path.join(gongsi_file, cell.value + ".docx")))

                try:
                    shutil.copy(os.path.join(gongsi_file, cell.value + ".docx"), lingshi_path)

                except Exception as e:
                    with open('error_log.txt', mode='at') as f:
                        f.write("\n"+cell.value + "---" + lingshi_path)
